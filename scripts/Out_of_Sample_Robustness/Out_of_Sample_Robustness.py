import sys
import os
import math
import datetime
import pickle
import time      
import seaborn as sns
import xlsxwriter
from math import *
from numpy import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')
global file_nameg  # file_path
file_nameg = os.getcwd() + '\\'


def write_xlsx(file_name, data):
    sheet_name = list(data.keys())
    workbook = xlsxwriter.Workbook(file_name)
    for s in range(len(sheet_name)):
        worksheet = workbook.add_worksheet(sheet_name[s])
        data_temp = np.array(data[sheet_name[s]])
        if len(data_temp.shape) == 0:
            worksheet.write(0, 0, data_temp)
        elif len(data_temp.shape) == 1:
            row = data_temp.shape[0]
            for i in range(row):
                worksheet.write(i, 0, data_temp[i])
        else:
            row = data_temp.shape[0]
            col = data_temp.shape[1]
            for i in range(row):
                for j in range(col):
                    worksheet.write(i, j, data_temp[i, j])

    workbook.close()


def read_xlsx(file_name):
    wb = load_workbook(file_name)
    sheet_name = wb.sheetnames

    if "Sheet1" in sheet_name:
        sheet_name.remove("Sheet1")

    Data = {}
    for name in sheet_name:
        sheet = wb[name]
        row = sheet.max_row
        col = sheet.max_column
        Data[name] = np.zeros((row, col))
        for i in range(row):
            for j in range(col):
                Data[name][i, j] = sheet.cell(row=(i + 1), column=(j + 1)).value

    return Data


def CFLP(y, d, f, I, J, c, r, q):
    model = gp.Model('D_CFLP')
    # variables
    x = model.addVars(I, J, lb=0, ub=+GRB.INFINITY, vtype=GRB.CONTINUOUS)
    # constraints

    for j in range(J):
        model.addConstr(x.sum('*', j) - d[j] <= 0)

    for i in range(I):
        model.addConstr(x.sum(i, '*') - q[i] * y[i] <= 0)

    obj = sum(f[i] * y[i]
              for i in range(I)) \
          + sum(r[j] * d[j]
                for j in range(J)) \
          + sum((c[i, j] - r[j]) * x[i, j]
                for i in range(I) for j in range(J))

    model.setObjective(obj, sense=GRB.MINIMIZE)
    model.setParam('OutputFlag', 0)
    model.optimize()
    solx = model.getAttr('x', x)
    solx_a = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            solx_a[i, j] = solx[i, j]
    objValue = model.objVal
    TSC = 0
    for i in range(I):
        for j in range(J):
            TSC += c[i, j] * solx_a[i, j]
    STC = 0
    for j in range(J):
        STC += r[j]*(d[j] - np.sum(solx_a[:, j]))
    return objValue, TSC, STC


if __name__ == '__main__':
    I = 10
    J = 20
    NN = 2000    # number of sample for one test set
    NN_S = 100   # number of test set

    y_SDRO = array([0, 1, 0, 1, 1, 0, 0, 1, 0, 0])

    y_DRO = array([0, 1, 0, 0, 1, 0, 0, 0, 1, 1])

    y_SAA = array([0, 1, 0, 0, 1, 0, 0, 1, 0, 0])

    Parameters_dict = read_xlsx(file_nameg + 'Parameters.xlsx')
    c = Parameters_dict['transportation_cost']
    r = Parameters_dict['penalty'].reshape(J)
    q = Parameters_dict['capacity'].reshape(I)
    f = Parameters_dict['fixed_cost'].reshape(I)

    phi = array([0, 1, 3, 5])

    d_sample_set_dict = {}
    for num_phi in range(4):
        key = 'phi_%d' % phi[num_phi]
        d_sample_set_dict[key] = read_xlsx(file_nameg + "d_sample_set_phi_"+str(int(phi[num_phi]))+".xlsx")

    d_sample_set_a_dict = {}

    for num_phi in range(4):
        key = 'phi_%d' % phi[num_phi]
        d_sample_set_a_dict[key] = np.zeros((NN, J, NN_S))
        for i in range(NN_S):
            set_key = 's%d' % i
            d_sample_set_a_dict[key][:, :, i] = d_sample_set_dict[key][set_key]

    for num_phi in range(4):
        key = 'phi_%d' % phi[num_phi]

        total_cost = np.zeros((NN, 3, NN_S))
        shortage_cost = np.zeros((NN, 3, NN_S))
        transportation_cost = np.zeros((NN, 3, NN_S))

        for i in range(NN_S):
            for nn in range(NN):
                TC_DRO, TSC_DRO, STC_DRO = CFLP(y_DRO, d_sample_set_a_dict[key][nn, :, i], f, I, J, c, r, q)
                TC_SDRO, TSC_SDRO, STC_SDRO = CFLP(y_SDRO, d_sample_set_a_dict[key][nn, :, i], f, I, J, c, r, q)
                TC_SAA, TSC_SAA, STC_SAA = CFLP(y_SAA, d_sample_set_a_dict[key][nn, :, i], f, I, J, c, r, q)
                total_cost[nn, 0, i] = TC_DRO
                total_cost[nn, 1, i] = TC_SDRO
                total_cost[nn, 2, i] = TC_SAA
                transportation_cost[nn, 0, i] = TSC_DRO
                transportation_cost[nn, 1, i] = TSC_SDRO
                transportation_cost[nn, 2, i] = TSC_SAA
                shortage_cost[nn, 0, i] = STC_DRO
                shortage_cost[nn, 1, i] = STC_SDRO
                shortage_cost[nn, 2, i] = STC_SAA

        total_cost_dict = {}
        shortage_cost_dict = {}
        transportation_cost_dict = {}

        total_cost_average = np.zeros((NN_S, 3))
        shortage_cost_average = np.zeros((NN_S, 3))
        transportation_cost_average = np.zeros((NN_S, 3))

        for i in range(NN_S):
            key_set = 'd%d' % i
            total_cost_dict[key_set] = total_cost[:, :, i]
            total_cost_average[i, :] = mean(total_cost[:, :, i], 0)

            transportation_cost_dict[key_set] = transportation_cost[:, :, i]
            transportation_cost_average[i, :] = mean(transportation_cost[:, :, i], 0)

            shortage_cost_dict[key_set] = shortage_cost[:, :, i]
            shortage_cost_average[i, :] = mean(shortage_cost[:, :, i], 0)

        write_xlsx(file_nameg + 'Ave_total_cost_robustness_phi_'+str(int(phi[num_phi]))+'.xlsx', {'obj': total_cost_average})
        write_xlsx(file_nameg + 'Ave_transportation_cost_robustness_phi_'+str(int(phi[num_phi]))+'.xlsx', {'obj': transportation_cost_average})
        write_xlsx(file_nameg + 'Ave_shortage_cost_robustness_phi_'+str(int(phi[num_phi]))+'.xlsx', {'obj': shortage_cost_average})

