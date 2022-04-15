import sys
import os
import math
import datetime
import pickle
import time
import xlsxwriter
from math import *
from numpy import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook
import warnings


def write_xlsx(file_name, data):
    sheet_name = list(data.keys())
    workbook = xlsxwriter.Workbook(file_name)
    for s in range(len(sheet_name)):
        worksheet = workbook.add_worksheet(sheet_name[s])
        data_temp = np.array(data[sheet_name[s]])
        if len(data_temp.shape) == 0:
            worksheet.write(0, 0, data_temp)
        elif len(data_temp.shape) == 1:
            row = data_temp.shape[0]
            for i in range(row):
                worksheet.write(i, 0, data_temp[i])
        else:
            row = data_temp.shape[0]
            col = data_temp.shape[1]
            for i in range(row):
                for j in range(col):
                    worksheet.write(i, j, data_temp[i, j])

    workbook.close()


def read_xlsx(file_name):
    wb = load_workbook(file_name)
    sheet_name = wb.sheetnames
    if "Sheet1" in sheet_name:
        sheet_name.remove("Sheet1")

    Data = {}
    for name in sheet_name:
        sheet = wb[name]
        row = sheet.max_row
        col = sheet.max_column
        Data[name] = np.zeros((row, col))
        for i in range(row):
            for j in range(col):
                Data[name][i, j] = sheet.cell(row=(i + 1), column=(j + 1)).value

    return Data


def find_EP(I, J, l, c, r):
    IEP = gp.Model("Initial Extreme Point")

    nu = IEP.addVars(J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="nu")

    lambda_IMP = IEP.addVars(I, lb=0, ub=+GRB.INFINITY,
                             vtype=GRB.CONTINUOUS, name="lambda")

    for i in range(I):
        for j in range(J):
            IEP.addConstr(-1 * nu[j] - lambda_IMP[i] - c[i, j] + r[j] <= 0,
                          "c_{}_{}".format(i, j))

    IEP.setObjective(0, sense=GRB.MINIMIZE)

    IEP.setParam("OutputFlag", 0)
    IEP.optimize()

    opti_nu = IEP.getAttr("X", nu).select("*")
    opti_lambda = IEP.getAttr("X", lambda_IMP).select("*")

    return opti_nu, opti_lambda


def ini_e_points(I, J, L, c, r):
    nu_hat = np.ones((J, 1))
    lambda_hat = np.ones((I, 1))

    for l in range(L):
        nu_temp, lambda_temp = find_EP(I, J, l, c, r)
        nu_hat = np.column_stack((nu_hat, nu_temp))
        lambda_hat = np.column_stack((lambda_hat, lambda_temp))
    nu_hat = nu_hat[:, 1:(L + 1)]
    lambda_hat = lambda_hat[:, 1:(L + 1)]

    return nu_hat, lambda_hat


def SP_MICP(I, J, K, M_k, 
            y_star, beta_u_k_star, beta_l_k_star,
            rho_k_star, A_k, b_k, Sigma_k, d_bar_k,
            c, r, q):

    # term 1
    q_bound = q.max()

    # term 2
    d_u = np.zeros((J, K))
    for j in range(J):
        for k in range(K):
            d_u_j_k = gp.Model("d_u_{}_{}".format(j, k))

            d_tmp = d_u_j_k.addVars(J, lb=0, ub=+GRB.INFINITY,
                                    vtype=GRB.CONTINUOUS, name="d")

            for m_k in range(M_k):
                c_tmp = gp.LinExpr(A_k[m_k, :], d_tmp.select("*"))
                c_tmp.addConstant(-1 * b_k[m_k])
                d_u_j_k.addConstr(c_tmp <= 0, "c_{}".format(m_k))

            d_u_j_k.setObjective(d_tmp[j], GRB.MAXIMIZE)

            d_u_j_k.setParam("OutputFlag", 0)
            d_u_j_k.optimize()

            d_u[j, k] = d_u_j_k.getAttr("ObjVal")
    d_bound = d_u.max()

    # term_3
    term_3_1 = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            tmp = 2 * (r[j] - c[i, j])
            if (tmp > 0):
                term_3_1[i, j] = tmp

    term_3_2 = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            tmp = r[j] - c[i, j]
            if (tmp < 0):
                term_3_2[i, j] = tmp

    r_bound = term_3_1.max() - term_3_2.min()
    r_bar = term_3_1.max()

    M = max(q_bound, d_bound, r_bound, r_bar)

    MICP = gp.Model("separation problem")
    
    # 声明变量
    d = MICP.addVars(J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="d")

    w = MICP.addVars(1, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="w")

    x = MICP.addVars(I, J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="x")

    nu = MICP.addVars(J, lb=0, ub=+GRB.INFINITY,
                      vtype=GRB.CONTINUOUS, name="nu")

    lambda_MICP = MICP.addVars(I, lb=0, ub=+GRB.INFINITY,
                               vtype=GRB.CONTINUOUS, name="lambda")

    pi_1 = MICP.addVars(J, vtype=GRB.BINARY, name="pi1")

    pi_2 = MICP.addVars(I, vtype=GRB.BINARY, name="pi2")

    pi_3 = MICP.addVars(I, J, vtype=GRB.BINARY, name="pi3")

    d_pro = MICP.addVars(J, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                         vtype=GRB.CONTINUOUS, name="d_pro")

    # c1
    for j in range(J):
        MICP.addConstr(x.sum("*", j) - d[j] <= 0,
                       "c1_{}".format(j))

    # c2
    for i in range(I):
        MICP.addConstr(x.sum(i, "*") - q[i] * y_star[i] <= 0,
                       "c2_{}".format(i))
    # c3
    for i in range(I):
        for j in range(J):
            MICP.addConstr(nu[j] + lambda_MICP[i] - r[j] + c[i, j] >= 0,
                           "c3_{}_{}".format(i, j))

    # c4
    for j in range(J):
        MICP.addConstr(d[j] - x.sum("*", j) - d_bound + d_bound * pi_1[j] <= 0,
                       "c4_{}".format(j))

    # c5
    for i in range(I):
        MICP.addConstr(x.sum(i, "*") - q[i] * y_star[i] + q_bound - q_bound * pi_2[i] >= 0,
                       "c5_{}".format(i))

    # c6
    for i in range(I):
        for j in range(J):
            MICP.addConstr(x[i, j] - q_bound + q_bound * pi_3[i, j] <= 0,
                           "c6_{}_{}".format(i, j))

    # c7
    for j in range(J):
        MICP.addConstr(nu[j] - r_bar * pi_1[j] <= 0, "c7_{}".format(j))

    # c8
    for i in range(I):
        MICP.addConstr(lambda_MICP[i] - r_bar * pi_2[i] <= 0, "c8_{}".format(i))

    # c9
    for i in range(I):
        for j in range(J):
            MICP.addConstr(nu[j] + lambda_MICP[i]
                           - r_bound * pi_3[i, j] - r[j] + c[i, j] <= 0,
                           "c9_{}_{}".format(i, j))

    # c10
    for m_k in range(M_k):
        c10 = gp.LinExpr(A_k[m_k, :], d.select("*"))
        c10.addConstant(-1 * b_k[m_k])
        MICP.addConstr(c10 <= 0, "c10_{}".format(m_k))

    # c11.0
    for j in range(J):
        c11_0 = gp.LinExpr(Sigma_k[j, :], d.select("*"))
        c11_0.addConstant(-1 * np.dot(Sigma_k[j, :], d_bar_k))
        c11_0.addTerms(-1, d_pro[j])
        MICP.addConstr(c11_0 == 0, "c11_0_{}".format(j))

    # c11
    c11 = gp.QuadExpr()
    for j in range(J):
        c11.add(d_pro[j] * d_pro[j], 1)
    c11.add(w[0] * w[0], -1)
    MICP.addQConstr(c11 <= 0, "c11")

    obj = gp.LinExpr()
    for i in range(I):
        obj.add(gp.LinExpr(c[i, :] - r, x.select(i, "*")), 1)

    obj.add(gp.LinExpr(r - beta_u_k_star + beta_l_k_star, d.select("*")), 1)

    obj.addTerms(-1 * rho_k_star, w[0])

    MICP.setObjective(obj, GRB.MAXIMIZE)

    MICP.setParam("OutputFlag", 0)
    MICP.optimize()

    run_time_MICP = MICP.getAttr('Runtime')
    
    Omega_k = MICP.getAttr("ObjVal")
    nu_k_hat = MICP.getAttr("X", nu).select("*")
    lambda_k_hat = MICP.getAttr("X", lambda_MICP).select("*")
    d_hat = MICP.getAttr("X", d).select("*")

    return Omega_k, nu_k_hat, lambda_k_hat, run_time_MICP, d_hat


def Q_G(I, J, L, M_k, 
        A_k, b_k, Sigma_k, d_bar_k, 
        d_uk, d_lk, sigma_k, 
        nu_hat, lambda_hat,
        q, r, y): 

    Q = gp.Model("Q_k(y)")

    zeta = Q.addVars(L, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="zeta")

    eta = Q.addVars(J, L, lb=0, ub=+GRB.INFINITY,
                    vtype=GRB.CONTINUOUS, name="eta")

    psi = Q.addVars(J, L, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                    vtype=GRB.CONTINUOUS, name="psi")

    gamma = Q.addVars(L, lb=0, ub=+GRB.INFINITY,
                      vtype=GRB.CONTINUOUS, name="gamma")

    # c1
    Q.addConstr(zeta.sum("*") == 1, "c1")

    # c2
    for j in range(J):
        Q.addConstr(eta.sum(j, "*") - d_uk[j] <= 0, "c2_{}".format(j))

    # c3
    for j in range(J):
        Q.addConstr(eta.sum(j, "*") - d_lk[j] >= 0, "c3_{}".format(j))

    # c4
    for e_k in range(L):
        for m_k in range(M_k):
            c4 = gp.LinExpr(A_k[m_k, :], eta.select("*", e_k))
            c4.addTerms(-1 * b_k[m_k], zeta[e_k])
            Q.addConstr(c4 <= 0, "c4_{}_{}".format(e_k, m_k))

    # c5
    term_1_coef = np.dot(Sigma_k, d_bar_k)
    for e_k in range(L):
        for j in range(J):
            c5 = gp.LinExpr(-1 * Sigma_k[j, :], eta.select("*", e_k))
            c5.addTerms(term_1_coef[j], zeta[e_k])
            c5.addTerms(1, psi[j, e_k])
            Q.addConstr(c5 == 0, "c5_{}_{}".format(e_k, j))

    # c6
    Q.addConstr(gamma.sum('*') - sigma_k <= 0, "c6")

    # c7
    for e_k in range(L):
        c7 = gp.QuadExpr()
        for j in range(J):
            c7.add(psi[j, e_k] * psi[j, e_k])
        c7.add(gamma[e_k] * gamma[e_k], -1)
        Q.addConstr(c7 <= 0, "c7_{}".format(e_k))

    term_1 = gp.LinExpr()
    for e_k in range(L):
        term_1.add(gp.LinExpr(r - nu_hat[:, e_k], eta.select("*", e_k)))

    q_mult_y = np.multiply(q, y)
    term_2 = gp.LinExpr()
    for e_k in range(L):
        term_2.addTerms(np.dot(q_mult_y, lambda_hat[:, e_k]), zeta[e_k])

    obj = term_1
    obj.add(term_2, -1)
    
    Q.setObjective(obj, sense=GRB.MAXIMIZE)

    Q.setParam("OutputFlag", 0)
    Q.optimize()

    run_time_Q = Q.getAttr('Runtime')

    Q_k = Q.getAttr("ObjVal")
    
    opti_zeta = Q.getAttr("X", zeta)
    opti_zeta = opti_zeta.select("*")
    
    opti_eta = Q.getAttr("X", eta)
    opti_eta = opti_eta.select("*")

    G_k = np.ones(I)
    for i in range(I):
        G_k[i] = -1 * q[i] * np.dot(lambda_hat[i, :], opti_zeta)

    return Q_k, G_k, run_time_Q


def fix_y_RMP(y_star_decompose, 
              I, J, K, L, M_k, 
              A_k, b_k, Sigma_k, d_bar_k, 
              d_uk, d_lk, sigma_k, 
              nu_hat, lambda_hat, 
              f, p, q, r): 

    FYRMP = gp.Model("Fix y Relaxed Master Problem")

    alpha = FYRMP.addVars(K, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                          vtype=GRB.CONTINUOUS, name="alpha")

    beta_u = FYRMP.addVars(J, K, lb=0, ub=+GRB.INFINITY,
                           vtype=GRB.CONTINUOUS, name="beta_u")

    beta_l = FYRMP.addVars(J, K, lb=0, ub=+GRB.INFINITY,
                           vtype=GRB.CONTINUOUS, name="beta_l")

    rho = FYRMP.addVars(K, lb=0, ub=+GRB.INFINITY,
                        vtype=GRB.CONTINUOUS, name="rho")

    xi = FYRMP.addVars(M_k, L, K, lb=0, ub=+GRB.INFINITY,
                       vtype=GRB.CONTINUOUS, name="xi")

    phi = FYRMP.addVars(J, L, K, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                        vtype=GRB.CONTINUOUS, name="phi")

    # c1
    for k in range(K):
        term_3_coef = np.dot(Sigma_k[:, :, k], d_bar_k[:, k])
        for e_k in range(L):
            # term 3
            c1 = gp.LinExpr(term_3_coef, phi.select("*", e_k, k))
            # term 2
            c1.add(gp.LinExpr(b_k[:, k], xi.select("*", e_k, k)), -1)
            # term 1
            c1.addTerms(1, alpha[k])
            # term 4
            q_mult_y = np.multiply(q, y_star_decompose)
            c1.addConstant(np.dot(q_mult_y, lambda_hat[:, e_k]))
            
            FYRMP.addConstr(c1 >= 0, "c1_{}_{}".format(k, e_k))

    # c2
    for k in range(K):
        for e_k in range(L):
            for j in range(J):
                # term 1
                c2 = gp.LinExpr(A_k[:, j, k], xi.select("*", e_k, k))
                # term 2
                c2.add(gp.LinExpr(Sigma_k[:, j, k], 
                                  phi.select("*", e_k, k)), -1)
                # term 3 & 4
                c2.addTerms(1, beta_u[j, k])
                c2.addTerms(-1, beta_l[j, k])
                # term 5 & 6
                c2.addConstant(nu_hat[j, e_k] - r[j])

                FYRMP.addConstr(c2 >= 0, "c2_{}_{}_{}".format(k, e_k, j))

    # c3
    for k in range(K):
        for e_k in range(L):
            c3 = gp.QuadExpr()
            for j in range(J):
                c3.add(phi[j, e_k, k] * phi[j, e_k, k], 1)
            c3.add(rho[k] * rho[k], -1)

            FYRMP.addConstr(c3 <= 0, "c3_{}_{}".format(k, e_k))

    obj = gp.LinExpr()
    for k in range(K):
        term_2 = gp.LinExpr(d_uk[:, k], beta_u.select("*", k))
        term_2.add(gp.LinExpr(d_lk[:, k], beta_l.select("*", k)), -1)
        term_2.addTerms(1, alpha[k])
        term_2.addTerms(sigma_k[k], rho[k])

        obj.add(term_2, p[k])
    obj.addConstant(np.dot(f.T, y_star_decompose))

    FYRMP.setObjective(obj, sense=GRB.MINIMIZE)

    FYRMP.setParam("OutputFlag", 0)
    FYRMP.optimize()

    run_time_FYRMP = FYRMP.getAttr('Runtime')
    
    obj_val = FYRMP.getAttr("ObjVal")
    alpha_star = FYRMP.getAttr("X", alpha)
    beta_u_star = FYRMP.getAttr("X", beta_u)
    beta_l_star = FYRMP.getAttr("X", beta_l)
    rho_star = FYRMP.getAttr("X", rho)

    return obj_val, alpha_star, beta_u_star, beta_l_star, rho_star, \
        run_time_FYRMP


def decompose_RMP(I, J, K, L, M_k,
                  A_k, b_k, Sigma_k, d_bar_k,
                  d_uk, d_lk, sigma_k,
                  nu_hat, lambda_hat,
                  f, p, q, r):
    
    time_begin = time.perf_counter()
    
    print("*" * 60)
    print("Solving Relaxed Master Problem start")

    N_max = 10000
    
    y_all = np.zeros((I, N_max))
    
    Q_all = np.zeros((K, N_max))
    G_all = np.zeros((I, K, N_max))
    
    UB_all = float("inf") * np.ones(N_max)
    LB_all = -float("inf") * np.ones(N_max)
    
    run_time_Q_all = np.zeros((N_max, K))
    run_time_LB_all = np.zeros(N_max)

    cycle_index = 0
    flag = True
    while (flag == True):
        for k in range(K):
            Q_all[k, cycle_index], G_all[:, k, cycle_index], \
                run_time_Q_all[cycle_index, k] \
                = Q_G(I, J, L, M_k,
                      A_k[:, :, k], b_k[:, k],
                      Sigma_k[:, :, k], d_bar_k[:, k],
                      d_uk[:, k], d_lk[:, k], sigma_k[k],
                      nu_hat, lambda_hat,
                      q, r, y_all[:, cycle_index])

        UB_tmp = np.dot(f.T, y_all[:, cycle_index])[0] + np.dot(p, Q_all[:, cycle_index])
        
        if (UB_tmp < UB_all.min()):
            UB_all[cycle_index] = UB_tmp
        else:
            UB_all[cycle_index] = UB_all.min()

        LB = gp.Model("LB")

        y_LB = LB.addVars(I, vtype=GRB.BINARY, name="y")
        
        theta_LB = LB.addVars(K, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                              vtype=GRB.CONTINUOUS, name="theta")

        for n in range(cycle_index + 1):
            for k in range(K):
                c_LB = gp.LinExpr(G_all[:, k, n], y_LB.select("*"))
                c_LB.addConstant(-1 * np.dot(G_all[:, k, n], y_all[:, n]))
                c_LB.addConstant(Q_all[k, n])
                c_LB.addTerms(-1, theta_LB[k])
                LB.addConstr(c_LB <= 0, "c_{}_{}".format(n, k))

        obj = gp.LinExpr(f, y_LB.select("*"))
        obj.add(gp.LinExpr(p, theta_LB.select("*")))
        LB.setObjective(obj, sense=GRB.MINIMIZE)

        LB.setParam("OutputFlag", 0)
        LB.optimize()

        run_time_LB_all[cycle_index] = LB.getAttr('Runtime')
        
        obj_val_LB = LB.getAttr("ObjVal")
        y_square = LB.getAttr("X", y_LB)

        LB_tmp = obj_val_LB
        if (LB_tmp > LB_all.max()):
            LB_all[cycle_index] = LB_tmp
        else:
            LB_all[cycle_index] = LB_all.max()


        if (UB_all[cycle_index] <= LB_all[cycle_index] + 0.0001): 
            if ((LB_all[cycle_index] == LB_all[cycle_index - 5]) and 
                (UB_all[cycle_index] == UB_all[cycle_index - 5])): 
                flag = False
                y_star_decompose = np.array(y_square.select("*"))
                y_star_dec_temp = np.zeros(I)
                for i in range(I):
                    y_star_dec_temp[i] = round(y_star_decompose[i])

        cycle_index = cycle_index + 1
        y_all[:, cycle_index] = np.array(y_square.select("*"))


    obj_val, alpha_star, beta_u_star, beta_l_star, rho_star, \
        run_time_FYRMP = fix_y_RMP(y_star_dec_temp, 
                                   I, J, K, L, M_k, 
                                   A_k, b_k, Sigma_k, d_bar_k, 
                                   d_uk, d_lk, sigma_k, 
                                   nu_hat, lambda_hat, 
                                   f, p, q, r)
        
    run_time_dec_RMP = sum(run_time_Q_all) + sum(run_time_LB_all) \
        + run_time_FYRMP
        
    time_end = time.perf_counter()
    
    print("Solving Relaxed Master Problem end")
    print("optimal_solution: {}".format(y_star_dec_temp))
    print("optimal_value: {}".format(obj_val))
    print("computation_time: {}".format(run_time_dec_RMP))
    print("*" * 60)
    print("*" * 60)

    return obj_val, y_star_decompose, \
        alpha_star, beta_u_star, beta_l_star, rho_star, \
            run_time_dec_RMP, cycle_index


def Q_LP(y, d, q, I, J, c, r):
    LP = gp.Model("Q_LP")

    nu = LP.addVars(J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="nu")

    lambda_LP = LP.addVars(I, lb=0, ub=+GRB.INFINITY,
                            vtype=GRB.CONTINUOUS, name="lambda")

    for i in range(I):
        for j in range(J):
            LP.addConstr(-1 * nu[j] - lambda_LP[i] - c[i, j] + r[j] <= 0,
                          "c_{}_{}".format(i, j))

    obj = gp.LinExpr()
    obj.add(gp.LinExpr(d, nu.select("*")), -1)
    obj.add(gp.LinExpr(np.multiply(y, q), lambda_LP.select("*")), -1)

    LP.setObjective(obj, sense=GRB.MAXIMIZE)

    LP.setParam("OutputFlag", 0)
    LP.optimize()

    Runtime_QLP = LP.getAttr('Runtime')
    LP_nu = LP.getAttr("X", nu).select("*")
    LP_lambda = LP.getAttr("X", lambda_LP).select("*")
    obj_LP = LP.getAttr("ObjVal")
    return LP_nu, LP_lambda, obj_LP, Runtime_QLP


def S_NBD_I10_J30(I, J, K, M_k, A_k, b_k, Sigma_k, d_bar_k, d_uk, d_lk, sigma_k, c, f, p, q, r, file_nameg):

    time_star = time.perf_counter()
    L = J

    nu_hat_0, lambda_hat_0 = ini_e_points(I, J, L, c, r)
    nu_hat = nu_hat_0
    lambda_hat = lambda_hat_0

    obj_val_all = np.zeros(1)

    y_star_all = np.zeros((I, 1))

    alpha_star_all = np.zeros((K, 1))

    Omega_all = np.zeros((K, 1))

    UB = np.zeros(1)

    LB = np.zeros(1)

    UB_all = np.zeros(1)

    flag_count_all = np.zeros(1)
    
    run_time_dec_RMP_all = np.zeros(1)
    run_time_MICP_all = np.zeros(1)
    time_loop_all = np.zeros(1)
    run_time_all = np.ones(1)
    runtime_QLP_all = np.ones(1)

    cycle_all = np.zeros(1)

    gap_all = np.ones(1)
    gap_prime_all = np.ones(1)

    flag = 1
    times = 0

    while (flag == 1):
        time_begin_inside = time.perf_counter()

        times = times + 1
        
        print()
        print("=" * 60)
        print("           S_NBD_Capacitated_I_{}_J_{}_K_{}:the {}th iteration".format(I, J, K, times))

        obj_val_dec, y_star_dec, alpha_star_dec, beta_u_star_dec, beta_l_star_dec, rho_star_dec, \
                run_time_dec_RMP, cycle_index \
                    = decompose_RMP(I, J, K, L, M_k, A_k, b_k, Sigma_k, d_bar_k,
                                    d_uk, d_lk, sigma_k, nu_hat, lambda_hat, 
                                    f, p, q, r)

        y_star_dec_temp = np.zeros(I)
        for i in range(I):
            y_star_dec_temp[i] = round(y_star_dec[i])

        Omega = np.ones(K)

        flag_count = 0

        nu_hat_new = np.ones((J, 1))
        lambda_hat_new = np.ones((I, 1))

        y_star = y_star_dec_temp
        if times == 1:
            y_star_all[:, 0] = y_star
        else:
            y_star_all = np.column_stack((y_star_all, y_star))
        if times==1:
            obj_val_all[0] = obj_val_dec
        else:
            obj_val_all = np.column_stack((obj_val_all, obj_val_dec))

        if times == 1:
            LB[0] = obj_val_dec
        else:
            LB = np.row_stack((LB, obj_val_dec))

        UB_1 = np.zeros(K)
        UB_2 = np.zeros(K)
        UB_3 = np.zeros(K)
        
        run_time_MICP = np.zeros(K)
        run_time_QLP = np.zeros(K)
        
        for k in range(K):
            beta_u_k_star = np.array(beta_u_star_dec.select("*", k))
            beta_l_k_star = np.array(beta_l_star_dec.select("*", k))
            rho_k_star = np.array(rho_star_dec.select(k))

            Omega_k, nu_k_hat, lambda_k_hat, run_time_MICP[k], d_k_hat \
                = SP_MICP(I, J, K, M_k, 
                          y_star, beta_u_k_star, 
                          beta_l_k_star, rho_k_star, 
                          A_k[:, :, k], b_k[:, k], 
                          Sigma_k[:, :, k], d_bar_k[:, k], 
                          c, r, q)
    
            Omega[k] = Omega_k

            if (alpha_star_dec[k] < (Omega[k] - 0.1)):

                flag_count = flag_count + 1

                d_k_hat_temp = np.zeros(J)
                for j in range(J):
                    d_k_hat_temp[j] = d_k_hat[j]

                nu_LP_k, lambda_LP_k, obj_LP_k, run_time_QLP[k] = Q_LP(y_star, d_k_hat_temp, q, I, J, c, r)

                nu_hat_new = np.column_stack((nu_hat_new, np.array(nu_LP_k)))
                lambda_hat_new = np.column_stack((lambda_hat_new, np.array(lambda_LP_k)))

            UB_1[k] = sum(beta_u_k_star[j] * d_uk[j, k] for j in range(J))
            UB_2[k] = sum(beta_l_k_star[j] * d_lk[j, k] for j in range(J))
            UB_3[k] = rho_k_star[0] * sigma_k[k]

        nu_hat_new = nu_hat_new[:, 1:(flag_count + 1)]
        lambda_hat_new = lambda_hat_new[:, 1:(flag_count + 1)]

        temp_UB = sum(f[i, 0] * y_star[i] for i in range(I)) \
                 + sum(p[k] * (Omega[k] + UB_1[k] - UB_2[k] + UB_3[k])
                        for k in range(K))

        if times == 1:
            UB_all[0] = temp_UB
            UB[0] = np.min(UB_all)
        else:
            UB_all = np.row_stack((UB_all, temp_UB))
            UB = np.row_stack((UB, np.min(UB_all)))

        flag_count_all = np.row_stack((flag_count_all, flag_count))

        alpha_star_all = np.column_stack((alpha_star_all,
                                          np.array(alpha_star_dec.select("*"))))
    
        Omega_all = np.column_stack((Omega_all, Omega))
    
        for s in range(K):
            alpha_star_all[s, times] = round(alpha_star_all[s, times], 5)
            Omega_all[s, times] = round(Omega_all[s, times], 5)
            
        if (times == 1):
            L = flag_count
            nu_hat = nu_hat_new
            lambda_hat = lambda_hat_new
        else:
            L = L + flag_count
            nu_hat = np.column_stack((nu_hat, nu_hat_new))
            lambda_hat = np.column_stack((lambda_hat, lambda_hat_new))
            
        time_end_inside = time.perf_counter()

        if times == 1:
            run_time_dec_RMP_all[0] = run_time_dec_RMP
        else:
            run_time_dec_RMP_all = np.row_stack((run_time_dec_RMP_all,
                                                run_time_dec_RMP))

        if times == 1:
            run_time_MICP_all[0] = np.sum(run_time_MICP)
        else:
            run_time_MICP_all = np.row_stack((run_time_MICP_all,
                                             np.sum(run_time_MICP)))

        if times == 1:
            runtime_QLP_all[0] = sum(run_time_QLP)
        else:
            runtime_QLP_all = np.row_stack((runtime_QLP_all, sum(run_time_QLP)))

        if times == 1:
            time_loop_all[0] = time_end_inside - time_begin_inside
        else:
            time_loop_all = np.row_stack((time_loop_all,
                                         time_end_inside - time_begin_inside))

        run_time_all_temp = np.sum(run_time_dec_RMP_all) + np.sum(run_time_MICP_all) + np.sum(runtime_QLP_all)
        if times == 1:
            run_time_all[0] = run_time_all_temp
        else:
            run_time_all = np.row_stack((run_time_all, run_time_all_temp))

        run_time_iter_temp = run_time_dec_RMP + sum(run_time_QLP) + np.sum(run_time_MICP)
        if times == 1:
            run_time_iter_all = run_time_iter_temp
        else:
            run_time_iter_all = np.row_stack((run_time_iter_all, run_time_iter_temp))

        if times == 1:
            cycle_all[0] = cycle_index
        else:
            cycle_all = np.row_stack((cycle_all, cycle_index))

        if times == 1:
            UB_temp = UB[0]
            LB_temp = LB[0]
        else:
            UB_temp = UB[times-1, 0]
            LB_temp = LB[times-1, 0]

        if times == 1:
            gap_all[0] = (UB_temp - LB_temp) / UB_temp
        else:
            gap_all = np.row_stack((gap_all, (UB_temp - LB_temp) / UB_temp))

        '''Print'''
        print("-" * 60)
        print("-" * 60)
        print('                Upper_Bound = {}'.format(UB_temp))
        print("-" * 60)
        print('                Lower_Bound = {}'.format(LB_temp))
        
        print("-" * 60)
        print('              The {}th computation time: {}'.format(times,
                                                  run_time_dec_RMP + sum(run_time_MICP) + sum(run_time_QLP)))

        print('              The {}th gap = {}'.format(times,
                                                  np.round(gap_all[(times-1)], 4)))

        print('              Total computation time {}'.format(run_time_all_temp))
        
        gap_prime = np.sum(p[k_n]*(Omega[k_n] - alpha_star_dec.select("*")[k_n]) for k_n in range(K)) / obj_val_dec
        if times == 1:
            gap_prime_all[0] = gap_prime
        else:
            gap_prime_all = np.row_stack((gap_prime_all, gap_prime))

        if (flag_count == 0):
            flag = 0
        else:
            flag = 1

        write_xlsx(file_nameg + 'Result_S_NBD_I_' + str(I) + '_J_' + str(J) + '_K_' + str(K) + '.xlsx',
                   {'Runtime': run_time_all,
                    'Runtime_each_iteration': run_time_iter_all,
                    'gap': gap_all,
                    'Upper_Bound': UB,
                    'Lower_Bound': LB})

    time_end = time.perf_counter()
    time_all = time_end - time_star

    run_time = sum(run_time_dec_RMP_all) \
               + sum(run_time_MICP_all) + sum(runtime_QLP_all)


    return obj_val_all, y_star_all, alpha_star_all, Omega_all, \
         UB, LB, UB_all, \
             run_time_dec_RMP_all, run_time_MICP_all, time_loop_all, \
                 cycle_all, gap_all, gap_prime_all, run_time_all, time_all, runtime_QLP_all



    