import sys
import os
import math
import datetime
import pickle
import time
import xlsxwriter
from math import *
from numpy import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

global file_nameg  # file_path

file_nameg = os.getcwd() + '\\'


def write_xlsx(file_name, data):
    sheet_name = list(data.keys())
    workbook = xlsxwriter.Workbook(file_name)
    for s in range(len(sheet_name)):
        worksheet = workbook.add_worksheet(sheet_name[s])
        data_temp = np.array(data[sheet_name[s]])
        if len(data_temp.shape) == 0:
            worksheet.write(0, 0, data_temp)
        elif len(data_temp.shape) == 1:
            row = data_temp.shape[0]
            for i in range(row):
                worksheet.write(i, 0, data_temp[i])
        else:
            row = data_temp.shape[0]
            col = data_temp.shape[1]
            for i in range(row):
                for j in range(col):
                    worksheet.write(i, j, data_temp[i, j])

    workbook.close()


def read_xlsx(file_name):
    wb = load_workbook(file_name)
    sheet_name = wb.sheetnames

    if "Sheet1" in sheet_name:
        sheet_name.remove("Sheet1")

    Data = {}
    for name in sheet_name:
        sheet = wb[name]
        row = sheet.max_row
        col = sheet.max_column
        Data[name] = np.zeros((row, col))
        for i in range(row):
            for j in range(col):
                Data[name][i, j] = sheet.cell(row=(i + 1), column=(j + 1)).value

    return Data


def CFLP(y, d, f, I, J, c, r, q):
    model = gp.Model('D_CFLP')
    # variables
    x = model.addVars(I, J, lb=0, ub=+GRB.INFINITY, vtype=GRB.CONTINUOUS)
    # constraints

    for j in range(J):
        model.addConstr(x.sum('*', j) - d[j] <= 0)

    for i in range(I):
        for j in range(J):
            model.addConstr(x.sum(i, '*') - q[i] * y[i] <= 0)

    obj = sum(f[i] * y[i]
              for i in range(I)) \
          + sum(r[j] * d[j]
                for j in range(J)) \
          + sum((c[i, j] - r[j]) * x[i, j]
                for i in range(I) for j in range(J))

    model.setObjective(obj, sense=GRB.MINIMIZE)
    model.setParam('OutputFlag', 0)
    model.optimize()
    solx = model.getAttr('x', x)
    solx_a = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            solx_a[i, j] = solx[i, j]
    objValue = model.objVal
    return objValue


if __name__ == '__main__':
    I = 10
    J = 20
    K = 4 # the number of state
    y_SDRO = array([0, 1, 0, 1, 1, 0, 0, 1, 0, 0])
    nominal_SDRO = 1516337
    y_DRO = array([0, 1, 0, 0, 1, 0, 0, 0, 1, 1])
    nominal_DRO = 1665817

    Parameters_dict = read_xlsx(file_nameg + 'Parameters.xlsx')
    c = Parameters_dict['transportation_cost']
    r = Parameters_dict['penalty'].reshape(J)
    q = Parameters_dict['capacity'].reshape(I)
    f = Parameters_dict['fixed_cost'].reshape(I)

    demand_2019_dict = read_xlsx(file_nameg + 'demand_observation_2019.xlsx')
    demand_2019 = demand_2019_dict['d']

    total_cost_SDRO = np.zeros(K)
    total_cost_DRO = np.zeros(K)
    conservativeness_SDRO = np.zeros(K)
    conservativeness_DRO = np.zeros(K)
    for k in range(K):
        total_cost_SDRO[k] = CFLP(y_SDRO, demand_2019[:, k], f, I, J, c, r, q)
        total_cost_DRO[k] = CFLP(y_DRO, demand_2019[:, k], f, I, J, c, r, q)
        conservativeness_SDRO[k] = nominal_SDRO - total_cost_SDRO[k]
        conservativeness_DRO[k] = nominal_DRO - total_cost_DRO[k]

    sheet_name = "Result_real_case_2019"
    workbook = xlsxwriter.Workbook(file_nameg + 'Result_real_case_2019.xlsx')
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.write(0, 0, 'Model')
    worksheet.write(1, 0, 'DRO')
    worksheet.write(2, 0, 'S-DRO')
    worksheet.write(0, 1, 'Expected total cost')
    worksheet.write(1, 1, np.mean(total_cost_DRO))
    worksheet.write(2, 1, np.mean(total_cost_SDRO))
    worksheet.write(0, 2, 'Nominal expected total cost')
    worksheet.write(1, 2, nominal_DRO)
    worksheet.write(2, 2, nominal_SDRO)
    for k in range(K):
        worksheet.write(0, (3+k), 'Total cost state'+str(k+1))
        worksheet.write(1, (3+k), total_cost_DRO[k])
        worksheet.write(2, (3+k), total_cost_SDRO[k])
    for k in range(K):
        worksheet.write(0, (7+k), 'Conservativeness state'+str(k+1))
        worksheet.write(1, (7 + k), conservativeness_DRO[k])
        worksheet.write(2, (7 + k), conservativeness_SDRO[k])
    workbook.close()
