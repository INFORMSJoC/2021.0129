import sys
import os
import math
import datetime
import pickle
import time
import xlsxwriter
from math import *
from numpy import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook
import warnings


def write_xlsx(file_name, data):
    sheet_name = list(data.keys())
    workbook = xlsxwriter.Workbook(file_name)
    for s in range(len(sheet_name)):
        worksheet = workbook.add_worksheet(sheet_name[s])
        data_temp = np.array(data[sheet_name[s]])
        if len(data_temp.shape) == 0:
            worksheet.write(0, 0, data_temp)
        elif len(data_temp.shape) == 1:
            row = data_temp.shape[0]
            for i in range(row):
                worksheet.write(i, 0, data_temp[i])
        else:
            row = data_temp.shape[0]
            col = data_temp.shape[1]
            for i in range(row):
                for j in range(col):
                    worksheet.write(i, j, data_temp[i, j])

    workbook.close()


def read_xlsx(file_name):
    wb = load_workbook(file_name)
    sheet_name = wb.sheetnames

    if "Sheet1" in sheet_name:
        sheet_name.remove("Sheet1")

    Data = {}
    for name in sheet_name:
        sheet = wb[name]
        row = sheet.max_row
        col = sheet.max_column
        Data[name] = np.zeros((row, col))
        for i in range(row):
            for j in range(col):
                Data[name][i, j] = sheet.cell(row=(i + 1), column=(j + 1)).value
    # 返回dict
    return Data

    
def find_EP(I, J, l, c, r):
    IEP = gp.Model("Initial Extreme Point")

    nu = IEP.addVars(J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="nu")

    lambda_IMP = IEP.addVars(I, lb=0, ub=+GRB.INFINITY,
                             vtype=GRB.CONTINUOUS, name="lambda")

    for i in range(I):
        for j in range(J):
            IEP.addConstr(-1 * nu[j] - lambda_IMP[i] - c[i, j] + r[j] <= 0,
                          "c_{}_{}".format(i, j))

    IEP.setObjective(0, sense=GRB.MINIMIZE)

    IEP.setParam("OutputFlag", 0)
    IEP.optimize()

    opti_nu = IEP.getAttr("X", nu).select("*")
    opti_lambda = IEP.getAttr("X", lambda_IMP).select("*")

    return opti_nu, opti_lambda


def ini_e_points(I, J, L, c, r):
    nu_hat = np.ones((J, 1))
    lambda_hat = np.ones((I, 1))

    for l in range(L):
        nu_temp, lambda_temp = find_EP(I, J, l, c, r)
        nu_hat = np.column_stack((nu_hat, nu_temp))
        lambda_hat = np.column_stack((lambda_hat, lambda_temp))
    nu_hat = nu_hat[:, 1:(L + 1)]
    lambda_hat = lambda_hat[:, 1:(L + 1)]

    return nu_hat, lambda_hat


def SP_MICP(I, J, K, M_k, 
            y_star, beta_u_k_star, beta_l_k_star,
            rho_k_star, A_k, b_k, Sigma_k, d_bar_k,
            c, r, q):
    # 求解big M
    # term 1
    max_1 = np.multiply(q, y_star).max()

    # term 2
    d_u = np.zeros((J, K))
    for j in range(J):
        for k in range(K):
            d_u_j_k = gp.Model("d_u_{}_{}".format(j, k))

            d_tmp = d_u_j_k.addVars(J, lb=0, ub=+GRB.INFINITY,
                                    vtype=GRB.CONTINUOUS, name="d")

            for m_k in range(M_k):
                c_tmp = gp.LinExpr(A_k[m_k, :], d_tmp.select("*"))
                c_tmp.addConstant(-1 * b_k[m_k])
                d_u_j_k.addConstr(c_tmp <= 0, "c_{}".format(m_k))

            d_u_j_k.setObjective(d_tmp[j], GRB.MAXIMIZE)

            d_u_j_k.setParam("OutputFlag", 0)
            d_u_j_k.optimize()

            d_u[j, k] = d_u_j_k.getAttr("ObjVal")
    max_2 = d_u.max()

    # term_3
    term_3_1 = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            tmp = 2 * (r[j] - c[i, j])
            if (tmp > 0):
                term_3_1[i, j] = tmp

    term_3_2 = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            tmp = r[j] - c[i, j]
            if (tmp < 0):
                term_3_2[i, j] = tmp

    term_3 = term_3_1.max() - term_3_2.min()
    
    M = max(max_1, max_2, term_3)
    
    # 建立模型: Mixed Integer Conic Optimization Problem 
    MICP = gp.Model("separation problem")
    
    # 声明变量
    d = MICP.addVars(J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="d")

    w = MICP.addVars(1, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="w")

    x = MICP.addVars(I, J, lb=0, ub=+GRB.INFINITY,
                     vtype=GRB.CONTINUOUS, name="x")

    nu = MICP.addVars(J, lb=0, ub=+GRB.INFINITY,
                      vtype=GRB.CONTINUOUS, name="nu")

    lambda_MICP = MICP.addVars(I, lb=0, ub=+GRB.INFINITY,
                               vtype=GRB.CONTINUOUS, name="lambda")

    pi_1 = MICP.addVars(J, vtype=GRB.BINARY, name="pi1")

    pi_2 = MICP.addVars(I, vtype=GRB.BINARY, name="pi2")

    pi_3 = MICP.addVars(I, J, vtype=GRB.BINARY, name="pi3")

    d_pro = MICP.addVars(J, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                         vtype=GRB.CONTINUOUS, name="d_pro")
    
    # 声明约束
    # c1
    for j in range(J):
        MICP.addConstr(x.sum("*", j) - d[j] <= 0,
                       "c1_{}".format(j))

    # c2
    for i in range(I):
        MICP.addConstr(x.sum(i, "*") - q[i] * y_star[i] <= 0,
                       "c2_{}".format(i))
    # c3
    for i in range(I):
        for j in range(J):
            MICP.addConstr(nu[j] + lambda_MICP[i] - r[j] + c[i, j] >= 0,
                           "c3_{}_{}".format(i, j))

    # c4
    for j in range(J):
        MICP.addConstr(d[j] - x.sum("*", j) - M + M * pi_1[j] <= 0,
                       "c4_{}".format(j))

    # c5
    for i in range(I):
        MICP.addConstr(x.sum(i, "*") - q[i] * y_star[i] + M - M * pi_2[i] >= 0,
                       "c5_{}".format(i))

    # c6
    for i in range(I):
        for j in range(J):
            MICP.addConstr(x[i, j] - M + M * pi_3[i, j] <= 0,
                           "c6_{}_{}".format(i, j))

    # c7
    for j in range(J):
        MICP.addConstr(nu[j] - M * pi_1[j] <= 0, "c7_{}".format(j))

    # c8
    for i in range(I):
        MICP.addConstr(lambda_MICP[i] - M * pi_2[i] <= 0, "c8_{}".format(i))

    # c9
    for i in range(I):
        for j in range(J):
            MICP.addConstr(nu[j] + lambda_MICP[i]
                           - M * pi_3[i, j] - r[j] + c[i, j] <= 0,
                           "c9_{}_{}".format(i, j))

    # c10
    for m_k in range(M_k):
        c10 = gp.LinExpr(A_k[m_k, :], d.select("*"))
        c10.addConstant(-1 * b_k[m_k])
        MICP.addConstr(c10 <= 0, "c10_{}".format(m_k))

    # c11.0
    for j in range(J):
        c11_0 = gp.LinExpr(Sigma_k[j, :], d.select("*"))
        c11_0.addConstant(-1 * np.dot(Sigma_k[j, :], d_bar_k))
        c11_0.addTerms(-1, d_pro[j])
        MICP.addConstr(c11_0 == 0, "c11_0_{}".format(j))

    # c11
    c11 = gp.QuadExpr()
    for j in range(J):
        c11.add(d_pro[j] * d_pro[j], 1)
    c11.add(w[0] * w[0], -1)
    MICP.addQConstr(c11 <= 0, "c11")

    obj = gp.LinExpr()
    for i in range(I):
        obj.add(gp.LinExpr(c[i, :] - r, x.select(i, "*")), 1)

    obj.add(gp.LinExpr(r - beta_u_k_star + beta_l_k_star, d.select("*")), 1)

    obj.addTerms(-1 * rho_k_star, w[0])
    
    # 声明目标函数
    MICP.setObjective(obj, GRB.MAXIMIZE)

    # 模型求解
    MICP.setParam("OutputFlag", 0)
    MICP.optimize()
    
    # 模型求解时间
    run_time_MICP = MICP.getAttr('Runtime')
    
    Omega_k = MICP.getAttr("ObjVal")
    nu_k_hat = MICP.getAttr("X", nu).select("*")
    lambda_k_hat = MICP.getAttr("X", lambda_MICP).select("*")

    return Omega_k, nu_k_hat, lambda_k_hat, run_time_MICP


def fix_y_RMP(y_star_decompose, 
              I, J, K, L, M_k, 
              A_k, b_k, Sigma_k, d_bar_k, 
              d_uk, d_lk, sigma_k, 
              nu_hat, lambda_hat, 
              f, p, q, r): 

    FYRMP = gp.Model("Fix y Relaxed Master Problem")

    alpha = FYRMP.addVars(K, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                          vtype=GRB.CONTINUOUS, name="alpha")

    beta_u = FYRMP.addVars(J, K, lb=0, ub=+GRB.INFINITY,
                           vtype=GRB.CONTINUOUS, name="beta_u")

    beta_l = FYRMP.addVars(J, K, lb=0, ub=+GRB.INFINITY,
                           vtype=GRB.CONTINUOUS, name="beta_l")

    rho = FYRMP.addVars(K, lb=0, ub=+GRB.INFINITY,
                        vtype=GRB.CONTINUOUS, name="rho")

    xi = FYRMP.addVars(M_k, L, K, lb=0, ub=+GRB.INFINITY,
                       vtype=GRB.CONTINUOUS, name="xi")

    phi = FYRMP.addVars(J, L, K, lb=-GRB.INFINITY, ub=+GRB.INFINITY,
                        vtype=GRB.CONTINUOUS, name="phi")

    # c1
    for k in range(K):
        term_3_coef = np.dot(Sigma_k[:, :, k], d_bar_k[:, k])
        for e_k in range(L):
            # term 3
            c1 = gp.LinExpr(term_3_coef, phi.select("*", e_k, k))
            # term 2
            c1.add(gp.LinExpr(b_k[:, k], xi.select("*", e_k, k)), -1)
            # term 1
            c1.addTerms(1, alpha[k])
            # term 4
            q_mult_y = np.multiply(q, y_star_decompose)
            c1.addConstant(np.dot(q_mult_y, lambda_hat[:, e_k]))
            
            FYRMP.addConstr(c1 >= 0, "c1_{}_{}".format(k, e_k))

    # c2
    for k in range(K):
        for e_k in range(L):
            for j in range(J):
                # term 1
                c2 = gp.LinExpr(A_k[:, j, k], xi.select("*", e_k, k))
                # term 2
                c2.add(gp.LinExpr(Sigma_k[:, j, k], 
                                  phi.select("*", e_k, k)), -1)
                # term 3 & 4
                c2.addTerms(1, beta_u[j, k])
                c2.addTerms(-1, beta_l[j, k])
                # term 5 & 6
                c2.addConstant(nu_hat[j, e_k] - r[j])

                FYRMP.addConstr(c2 >= 0, "c2_{}_{}_{}".format(k, e_k, j))

    # c3
    for k in range(K):
        for e_k in range(L):
            c3 = gp.QuadExpr()
            for j in range(J):
                c3.add(phi[j, e_k, k] * phi[j, e_k, k], 1)
            c3.add(rho[k] * rho[k], -1)

            FYRMP.addConstr(c3 <= 0, "c3_{}_{}".format(k, e_k))

    obj = gp.LinExpr()
    for k in range(K):
        term_2 = gp.LinExpr(d_uk[:, k], beta_u.select("*", k))
        term_2.add(gp.LinExpr(d_lk[:, k], beta_l.select("*", k)), -1)
        term_2.addTerms(1, alpha[k])
        term_2.addTerms(sigma_k[k], rho[k])

        obj.add(term_2, p[k])
    obj.addConstant(np.dot(f.T, y_star_decompose))

    FYRMP.setObjective(obj, sense=GRB.MINIMIZE)

    FYRMP.setParam("OutputFlag", 0)
    FYRMP.optimize()

    run_time_FYRMP = FYRMP.getAttr('Runtime')
    
    obj_val = FYRMP.getAttr("ObjVal")
    alpha_star = FYRMP.getAttr("X", alpha)
    beta_u_star = FYRMP.getAttr("X", beta_u)
    beta_l_star = FYRMP.getAttr("X", beta_l)
    rho_star = FYRMP.getAttr("X", rho)

    return obj_val, alpha_star, beta_u_star, beta_l_star, rho_star, \
        run_time_FYRMP


def CLP_EW_EXACT_fixy(y_star, I, J, K, M_k, 
                      A_k, b_k, Sigma_k, d_bar_k, 
                      d_uk, d_lk, sigma_k, 
                      c, f, p, q, r): 
    L = 20
    
    nu_hat, lambda_hat = ini_e_points(I, J, L, c, r)
    
    obj_val_all = np.zeros(1)
    alpha_star_all = np.zeros((K, 1))
    Omega_all = np.zeros((K, 1))
    UB = np.zeros(1)
    LB = np.zeros(1)
    UB_all = np.zeros(1)
    flag_count_a = np.zeros(1)
    
    flag = 1
    times = 0
    UB_temp = 1000
    LB_temp = 100
    while (flag == 1 and times <= 1000):
        times = times + 1
        
        obj_val, alpha_star, beta_u_star, beta_l_star, rho_star, \
            run_time_FYRMP \
                = fix_y_RMP(y_star, 
                            I, J, K, L, M_k, 
                            A_k, b_k, Sigma_k, d_bar_k, 
                            d_uk, d_lk, sigma_k, 
                            nu_hat, lambda_hat, 
                            f, p, q, r)
            
        obj_val_all = np.column_stack((obj_val_all, obj_val))
        
        if times == 1:
            LB[0] = obj_val
        else:
            LB = np.column_stack((LB, obj_val))
    
        Omega = np.ones(K)
        nu_hat_new = np.ones((J, 1))
        lambda_hat_new = np.ones((I, 1))
    
        flag_count = 0
    
        temp_1 = np.zeros(K)
        temp_2 = np.zeros(K)
        temp_3 = np.zeros(K)
        for k in range(K):
            beta_u_k_star = np.array(beta_u_star.select("*", k))
            beta_l_k_star = np.array(beta_l_star.select("*", k))
            rho_k_star = np.array(rho_star.select(k))
    
            Omega_k, nu_k_hat, lambda_k_hat, run_time_MICP \
                = SP_MICP(I, J, K, M_k, 
                          y_star, beta_u_k_star, 
                          beta_l_k_star, rho_k_star, 
                          A_k[:, :, k], b_k[:, k], 
                          Sigma_k[:, :, k], d_bar_k[:, k], 
                          c, r, q)
        
            
            Omega[k] = Omega_k
    
            if (alpha_star[k] < (Omega[k]-1)):
                flag_count = flag_count + 1
                nu_hat_new = np.column_stack((nu_hat_new, np.array(nu_k_hat)))
                lambda_hat_new = np.column_stack((lambda_hat_new, np.array(lambda_k_hat)))
    
            temp_1[k] = sum(beta_u_k_star[j] * d_uk[j, k] for j in range(J))
            temp_2[k] = sum(beta_l_k_star[j] * d_lk[j, k] for j in range(J))
            temp_3[k] = rho_k_star[0] * sigma_k[k]
    
        nu_hat_new = nu_hat_new[:, 1:(flag_count + 1)]
        lambda_hat_new = lambda_hat_new[:, 1:(flag_count + 1)]
    
        beta_u_star_a = np.array(beta_u_star.select('*', '*'))
        temp_UB = sum(f[i, 0] * y_star[i] for i in range(I)) \
                 + sum(p[k] * (Omega[k] + temp_1[k] - temp_2[k] + temp_3[k])
                        for k in range(K))
        if times == 1:
            UB_all[0] = temp_UB
            UB[0] = np.min(UB_all)
        else:
            UB_all = np.column_stack((UB_all, temp_UB))
            UB = np.column_stack((UB, np.min(UB_all)))
    
    
        alpha_star_all = np.column_stack((alpha_star_all,
                                          np.array(alpha_star.select("*"))))
    
        Omega_all = np.column_stack((Omega_all, Omega))
    
        for s in range(K):
            alpha_star_all[s, times] = round(alpha_star_all[s, times], 5)
            Omega_all[s, times] = round(Omega_all[s, times], 5)

        flag_count_a = np.column_stack((flag_count_a, flag_count))
    
        if times == 1:
            UB_temp = UB[0]
            LB_temp = LB[0]
        else:
            UB_temp = UB[0, times-1]
            LB_temp = LB[0, times-1]
        
        if (times == 1):
            L = K
            nu_hat = nu_hat_new
            lambda_hat = lambda_hat_new
        else:
            L = L + flag_count
            nu_hat = np.column_stack((nu_hat, nu_hat_new))
            lambda_hat = np.column_stack((lambda_hat, lambda_hat_new))
    
        if (flag_count == 0):
            flag = 0
        else:
            flag = 1
        
    return obj_val
    