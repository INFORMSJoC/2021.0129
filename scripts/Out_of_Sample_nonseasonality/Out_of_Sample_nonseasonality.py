import sys
import os
import math
import datetime
import pickle
import time      
import seaborn as sns
import xlsxwriter
from math import *
from numpy import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')

global file_nameg  # file path
file_nameg = os.getcwd() + '\\'


def write_xlsx(file_name, data):

    sheet_name = list(data.keys())
    workbook = xlsxwriter.Workbook(file_name)
    for s in range(len(sheet_name)):
        worksheet = workbook.add_worksheet(sheet_name[s])
        data_temp = np.array(data[sheet_name[s]])
        if len(data_temp.shape) == 0:
            worksheet.write(0, 0, data_temp)
        elif len(data_temp.shape) == 1:
            row = data_temp.shape[0]
            for i in range(row):
                worksheet.write(i, 0, data_temp[i])
        else:
            row = data_temp.shape[0]
            col = data_temp.shape[1]
            for i in range(row):
                for j in range(col):
                    worksheet.write(i, j, data_temp[i, j])

    workbook.close()


def read_xlsx(file_name):

    wb = load_workbook(file_name)
    sheet_name = wb.sheetnames

    if "Sheet1" in sheet_name:
        sheet_name.remove("Sheet1")

    Data = {}
    for name in sheet_name:
        sheet = wb[name]
        row = sheet.max_row
        col = sheet.max_column
        Data[name] = np.zeros((row, col))
        for i in range(row):
            for j in range(col):
                Data[name][i, j] = sheet.cell(row=(i + 1), column=(j + 1)).value

    return Data


def CFLP(y, d, f, I, J, c, r, q):
    model = gp.Model('D_CFLP')
    # variables
    x = model.addVars(I, J, lb=0, ub=+GRB.INFINITY, vtype=GRB.CONTINUOUS)
    # constraints

    for j in range(J):
        model.addConstr(x.sum('*', j) - d[j] <= 0)

    for i in range(I):
        model.addConstr(x.sum(i, '*') - q[i] * y[i] <= 0)

    obj = sum(f[i] * y[i]
              for i in range(I)) \
          + sum(r[j] * d[j]
                for j in range(J)) \
          + sum((c[i, j] - r[j]) * x[i, j]
                for i in range(I) for j in range(J))

    model.setObjective(obj, sense=GRB.MINIMIZE)
    model.setParam('OutputFlag', 0)
    model.optimize()
    solx = model.getAttr('x', x)
    solx_a = np.zeros((I, J))
    for i in range(I):
        for j in range(J):
            solx_a[i, j] = solx[i, j]
    objValue = model.objVal
    TSC = 0
    for i in range(I):
        for j in range(J):
            TSC += c[i, j] * solx_a[i, j]
    return objValue, TSC


if __name__ == '__main__':
    I = 10
    J = 20
    NN = 2000    # number of sample for one test set
    NN_S = 100   # number of test set

    y_SDRO = array([0, 1, 0, 1, 1, 0, 0, 1, 0, 0])
    nominal_SDRO = 1516337
    y_DRO = array([0, 1, 0, 0, 1, 0, 0, 0, 1, 1])
    nominal_DRO = 1665817

    Parameters_dict = read_xlsx(file_nameg + 'Parameters.xlsx')
    c = Parameters_dict['transportation_cost']
    r = Parameters_dict['penalty'].reshape(J)
    q = Parameters_dict['capacity'].reshape(I)
    f = Parameters_dict['fixed_cost'].reshape(I)

    d_sample_set = read_xlsx(file_nameg + "d_sample_set_nonseasonality.xlsx")

    d_sample_set_a = np.zeros((NN, J, NN_S))

    for i in range(NN_S):
        set_key = 's%d' % i
        d_sample_set_a[:, :, i] = d_sample_set[set_key]

    total_cost = np.zeros((NN, 2, NN_S))
    transportation_cost = np.zeros((NN, 2, NN_S))

    for i in range(NN_S):
        for nn in range(NN):
            TC_SDRO, TSC_SDRO= CFLP(y_SDRO, d_sample_set_a[nn, :, i], f, I, J, c, r, q)
            TC_DRO, TSC_DRO = CFLP(y_DRO, d_sample_set_a[nn, :, i], f, I, J, c, r, q)
            total_cost[nn, 0, i] = TC_SDRO
            total_cost[nn, 1, i] = TC_DRO
            transportation_cost[nn, 0, i] = TSC_SDRO
            transportation_cost[nn, 1, i] = TSC_DRO

    total_cost_dict = {}
    transportation_cost_dict = {}

    total_cost_average = np.zeros((NN_S, 2))
    transportation_cost_average = np.zeros((NN_S, 2))


    for i in range(NN_S):
        key = 'd%d' % i
        total_cost_dict[key] = total_cost[:, :, i]
        total_cost_average[i, :] = mean(total_cost[:, :, i], 0)

        transportation_cost_dict[key] = transportation_cost[:, :, i]
        transportation_cost_average[i, :] = mean(transportation_cost[:, :, i], 0)

    difference_total_cost_average = total_cost_average[:, 1] - total_cost_average[:, 0]
    difference_transportation_cost_average = transportation_cost_average[:, 1] - transportation_cost_average[:, 0]


    write_xlsx(file_nameg + 'Ave_total_cost_nonseasonality.xlsx', {'obj': total_cost_average})
    write_xlsx(file_nameg + 'Ave_transportation_cost_nonseasonality.xlsx', {'obj': transportation_cost_average})

    write_xlsx(file_nameg + 'Diff_Ave_total_cost_nonseasonality.xlsx', {'obj': difference_total_cost_average})
    write_xlsx(file_nameg + 'Diff_Ave_transportation_cost_nonseasonality.xlsx', {'obj': difference_transportation_cost_average})


